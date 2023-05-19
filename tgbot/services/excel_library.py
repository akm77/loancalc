from dataclasses import dataclass
from pprint import pprint
from tempfile import NamedTemporaryFile
from typing import List

from openpyxl.styles import Alignment, Font, Side, Border, PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.worksheet import Worksheet

EMPTY_ROW = [""]
CONDITION_OFFSET = ["", "", "", ""]
HEADER = CONDITION_OFFSET + ["Месяц", "Остаток ОД", "ОД", "%%", "ДП"]

THIN = Side(border_style="thin", color="000000")
DOTTED = Side(border_style="dotted", color="000000")
FINANCIAL = r'_-* #,##0.00 "₽"_-;-* #,##0.00 "₽"_-;_-* "-"?? "₽"_-;_-@_-'
INTEGER = '#,##0'


@dataclass
class LoanCondition:
    amount: int
    loan_period: int
    monthly_fee: float
    interest_rate: int


class PaymentSchedule:
    def __init__(self, condition: LoanCondition, data: List[List]):
        self.condition = condition
        self.data = data
        self.wb = None
        self.ws: Worksheet | None = None

    def _create_workbook(self):
        self.wb = Workbook()
        self.ws = self.wb.active

    def _fill_condition(self):
        self.ws["B2"] = "Сумма, руб."
        self.ws["C2"] = self.condition.amount
        self.ws["C2"].alignment = Alignment(horizontal='center',
                                            vertical='center')
        self.ws["C2"].font = Font(bold=True)
        self.ws["C2"].border = Border(top=DOTTED, left=DOTTED, right=DOTTED, bottom=DOTTED)
        self.ws["C2"].fill = PatternFill("solid", fgColor="00FFFFCC")
        self.ws["C2"].number_format = "# ##0"

        interest_row = 2
        if self.condition.interest_rate:
            self.ws["B3"] = "Процент годовой, %"
            self.ws["C3"] = self.condition.interest_rate
            self.ws["C3"].alignment = Alignment(horizontal='center',
                                                vertical='center')
            self.ws["C3"].font = Font(bold=True)
            self.ws["C3"].border = Border(top=DOTTED, left=DOTTED, right=DOTTED, bottom=DOTTED)
            self.ws["C3"].fill = PatternFill("solid", fgColor="00FFFFCC")
            self.ws["C3"].number_format = INTEGER
            interest_row = 3

        self.ws[f"B{interest_row + 1}"] = "Срок, мес"
        self.ws[f"C{interest_row + 1}"] = self.condition.loan_period
        self.ws[f"C{interest_row + 1}"].alignment = Alignment(horizontal='center',
                                                              vertical='center')
        self.ws[f"C{interest_row + 1}"].font = Font(bold=True)
        self.ws[f"C{interest_row + 1}"].border = Border(top=DOTTED, left=DOTTED, right=DOTTED, bottom=DOTTED)
        self.ws[f"C{interest_row + 1}"].fill = PatternFill("solid", fgColor="00FFFFCC")
        self.ws[f"C{interest_row + 1}"].number_format = "# ##0"
        self.ws[f"B{interest_row + 2}"] = "Размер аннуитетного платежа, руб"
        self.ws[f"C{interest_row + 2}"] = self.condition.monthly_fee
        self.ws[f"C{interest_row + 2}"].alignment = Alignment(horizontal='center',
                                                              vertical='center')
        self.ws[f"C{interest_row + 2}"].font = Font(bold=True)
        self.ws[f"C{interest_row + 2}"].border = Border(top=DOTTED, left=DOTTED, right=DOTTED, bottom=DOTTED)
        self.ws[f"C{interest_row + 2}"].number_format = FINANCIAL

    def _set_column_width(self):
        self.ws.column_dimensions["A"].width = 3
        self.ws.column_dimensions["B"].width = 33
        self.ws.column_dimensions["C"].width = 15
        self.ws.column_dimensions["D"].width = 3
        self.ws.column_dimensions["E"].width = 10
        self.ws.column_dimensions["F"].width = 15
        self.ws.column_dimensions["G"].width = 15
        self.ws.column_dimensions["H"].width = 15
        self.ws.column_dimensions["I"].width = 15

    def _format_range(self, cell_ranges, bold=True, number_format=None):
        for cell_range in cell_ranges:
            for cell in cell_range:
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center')
                if bold:
                    cell.font = Font(bold=True)
                cell.border = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
                if number_format:
                    cell.number_format = number_format

    def _format_data(self):
        self._format_range(self.ws["E2":"I2"])
        self._format_range(self.ws["E3": f"E{self.ws.max_row}"], bold=False, number_format=INTEGER)
        self._format_range(self.ws["F3": f"I{self.ws.max_row}"], bold=False, number_format=FINANCIAL)

    def _fill_worksheet(self):
        self.ws.append(EMPTY_ROW)
        self.ws.append(HEADER)
        _ = [self.ws.append(CONDITION_OFFSET + row) for row in self.data]
        self._set_column_width()
        self._fill_condition()
        self._format_data()

    def get_workbook(self):
        self._create_workbook()
        self._fill_worksheet()
        with NamedTemporaryFile() as tmp:
            self.wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        return stream
